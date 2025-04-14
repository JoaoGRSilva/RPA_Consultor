from openpyxl import load_workbook
import pandas as pd
import logging
from config import CONFIG

class ExcelProcessor:
    """Classe para processar planilhas relacionadas aos contratos."""
    
    @staticmethod
    def ler_contratos_pendentes(limite=None, modo_teste=False):
        """
        Lê a planilha de contratos e filtra os pendentes.
        
        Args:
            limite: Limite de contratos a processar
            modo_teste: Se True, processa apenas o primeiro contrato
            
        Returns:
            list: Lista de números de contrato a processar
        """
        try:
            caminho_excel = CONFIG['EXCEL_ESTEIRA']
            print(f"Lendo planilha de contratos: {caminho_excel}")
            
            df = pd.read_excel(caminho_excel, sheet_name='CONSOLIDADO', header=1)
            
            # Filtrar contratos pendentes
            contratos_vigentes = df[(df['STATUS'] == 'Vigente') & 
                                   (df['Status Omni'] == 'Demandado Para Logística') & 
                                   (df['CADASTRADO PLUXXE?'] == 'NÃO')]
                                   
            contratos_lista = contratos_vigentes['Numeração'].tolist()
            
            total = len(contratos_lista)
            print(f"Total de {total} contratos pendentes encontrados")
            
            # Aplicar limites se necessário
            if modo_teste:
                contratos_lista = contratos_lista[:1]
                logging.info("Modo de teste ativado - processando apenas 1 contrato")
            elif limite and limite > 0:
                contratos_lista = contratos_lista[:limite]
                logging.info(f"Limite aplicado - processando {len(contratos_lista)} contratos")
                
            return contratos_lista
            
        except Exception as e:
            logging.error(f"Erro ao ler planilha de contratos: {e}")
            return []
    
    @staticmethod
    def preencher_planilha(dados, linha_inicial=7):
        """
        Preenche a planilha PLUXXE com os dados do contrato.
        
        Args:
            dados: Dicionário com dados a inserir
            linha_inicial: Linha inicial para buscar espaço vazio
            
        Returns:
            bool: True se sucesso, False caso contrário
        """
        try:
            pluxxe_path = CONFIG['EXCEL_PLUXXE']
            wb = load_workbook(pluxxe_path)
            sheet = wb['Dados dos Beneficiários']

            # Encontrar próxima linha vazia
            linha_atual = linha_inicial
            while sheet.cell(row=linha_atual, column=1).value is not None:
                linha_atual += 1

            # Inserir dados
            for coluna, (chave, valor) in enumerate(dados.items(), start=1):
                sheet.cell(row=linha_atual, column=coluna, value=valor)

            wb.save(pluxxe_path)
            print(f"Dados inseridos na linha {linha_atual} da planilha PLUXXE")
            return True

        except Exception as e:
            logging.error(f"Erro ao preencher a planilha: {e}")
            return False