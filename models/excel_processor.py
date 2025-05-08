from openpyxl import load_workbook, Workbook
import pandas as pd
from datetime import date, datetime, timedelta
import os, re, shutil
from config import CONFIG
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from copy import copy


class ExcelProcessor:
    @staticmethod
    def ler_contratos_pendentes(limite=None, modo_teste=False):
        try:
            caminho_excel = CONFIG['EXCEL_ESTEIRA']
            print(f"Lendo planilha de contratos: {caminho_excel}")
            df = pd.read_excel(caminho_excel, sheet_name='CONSOLIDADO', header=1)

            contratos_vigentes = df[
                (df['STATUS'] == 'Vigente') &
                (df['CADASTRADO PLUXXE?'] == 'NÃO') &
                (df['Processado RPA'] == 'NÃO')
            ]

            contratos_lista = contratos_vigentes['Numeração'].tolist()
            total = len(contratos_lista)
            print(f"Total de {total} contratos pendentes encontrados")

            if modo_teste:
                contratos_lista = contratos_lista[:1]
                print("Modo de teste ativado - processando apenas 1 contrato")
            elif limite and limite > 0:
                contratos_lista = contratos_lista[:limite]
                print(f"Limite aplicado - processando {len(contratos_lista)} contratos")

            return contratos_lista

        except Exception as e:
            print(f"Erro ao ler planilha de contratos: {e}")
            return None

    @staticmethod
    def atualizar_esteira(resultados, caminho_excel):
        try:
            df_novos = pd.DataFrame(resultados)
            df_novos.rename(columns={'numero': 'Contrato', 'status': 'Status'}, inplace=True)

            if not os.path.exists(caminho_excel):
                df_novos.to_excel(
                    caminho_excel,
                    engine='openpyxl',
                    sheet_name='#RPA',
                    index=False
                )
                return True

            try:
                df_existente = pd.read_excel(caminho_excel, sheet_name='#RPA')
                df_final = pd.concat([df_existente, df_novos]).drop_duplicates(subset=['Contrato'], keep='last')

                colunas = ['Contrato', 'Status']
                if 'erro' in df_final.columns:
                    colunas.append('erro')

                with pd.ExcelWriter(caminho_excel, engine='openpyxl', mode='w') as writer:
                    df_final[colunas].to_excel(writer, sheet_name='#RPA', index=False)

                print(f"Planilha atualizada: {len(df_novos)} contratos")
                return True

            except Exception as e:
                colunas = ['Contrato', 'Status']
                if 'erro' in df_novos.columns:
                    colunas.append('erro')

                df_novos[colunas].to_excel(
                    caminho_excel,
                    engine='openpyxl',
                    sheet_name='#RPA',
                    index=False
                )
                return True

        except Exception as e:
            print(f"Falha crítica na atualização: {str(e)}")
            if "File is not a zip file" in str(e):
                os.remove(caminho_excel)
                pd.DataFrame({'Contrato': [], 'Status': []}).to_excel(
                    caminho_excel,
                    engine='openpyxl',
                    sheet_name='#RPA',
                    index=False
                )
            return False

    @staticmethod
    def preencher_planilha(dados, linha_inicial=7):
        try:
            pluxxe_path = 'planilha.xlsx'
            wb = load_workbook(pluxxe_path)
            sheet = wb['Dados dos Beneficiários']

            linha_atual = linha_inicial
            while sheet.cell(row=linha_atual, column=1).value is not None:
                linha_atual += 1

            for coluna, (chave, valor) in enumerate(dados.items(), start=1):
                sheet.cell(row=linha_atual, column=coluna, value=valor)

            wb.save(pluxxe_path)
            print(f"Dados inseridos na linha {linha_atual} da planilha PLUXXE")
            return True

        except Exception as e:
            print(f"Erro ao preencher a planilha: {e}")
            return False
    
    
    @staticmethod
    def limpar_planilha():
        try:
            shutil.copy(CONFIG['EXCEL_PLUXXE'], 'planilha.xlsx')
            pluxxe_path = 'planilha.xlsx'
            wb = load_workbook(pluxxe_path)
            ws = wb.active

            for linha in range(ws.max_row, 7, -1): 
                for celula in ws[linha]:
                    celula.value = None

            wb.save(pluxxe_path)
            print("Planilha limpa com sucesso!")

        except Exception as e:
            print(f"Erro ao limpar planilha: {e}")

    @staticmethod
    def renomear_saida():
        hoje = date.today()
        try:
            pluxxe_path = 'planilha.xlsx'
            dia_formatado = hoje.strftime("%d%m%y")
            nome_arquivo = f"PLANSIP4C_3230687_{dia_formatado}.xlsx"
            novo_caminho = os.path.join(os.path.dirname(pluxxe_path), nome_arquivo)
            os.rename(pluxxe_path, novo_caminho)
            print(f"Planilha renomeada para: {nome_arquivo}")
        except Exception as e:
            print(f"Erro ao renomear a planilha: {e}")

        wb = load_workbook(nome_arquivo)
        ws = wb.active
        img = Image("logo_pluxxe.png")
        ws.add_image(img, "A1")

        wb.save(nome_arquivo)


    @staticmethod
    def compilar_arquivos():
        # Determinando os dias anteriores
        hoje = datetime.today()
        dias_ate_segunda = hoje.weekday()
        ultima_segunda = hoje - timedelta(days=dias_ate_segunda + 7)
        ultima_sexta = hoje - timedelta(days = 4)

        # Loop para pegar os arquivos
        datas = []
        dia_atual = ultima_segunda
        while dia_atual <= ultima_sexta:
            datas.append(dia_atual.strftime("%d%m%y"))
            dia_atual += timedelta(days = 1)

        # Verifica a pasta pluxxe
        all_files = os.listdir(CONFIG['PLUXXE_FOLDER'])

        # Filtro dos dias anteriores
        arquivos_processados = []

        # Filtro
        padrao = r"PLANSIP4C_\d+_(\d{6})(?:\s*-\s*\d+)?"

        for arquivo in all_files:
            match = re.search(padrao, arquivo)
            if match and match.group(1) in datas:
                arquivos_processados.append(arquivo)
        
        print(arquivos_processados)

        return arquivos_processados


    @staticmethod
    def compilar_planilhas(arquivos, linha_inicio=8):
        wb_compilado = Workbook()
        ws_destino = wb_compilado.active
        ws_destino.title = 'Dados dos Beneficiários'

        primeira = True

        for arquivo in arquivos:
            # Adicionar o caminho completo ao nome do arquivo
            caminho_completo = os.path.join(CONFIG['PLUXXE_FOLDER'], arquivo)
            
            wb = load_workbook(caminho_completo, data_only=True)
            ws = wb.active

            if primeira:
                # Copiar configurações da planilha
                ws_destino.page_setup = ws.page_setup
                ws_destino.page_margins = ws.page_margins
                ws_destino.sheet_properties = ws.sheet_properties
                
                # Copiar larguras das colunas
                for i, column in enumerate(ws.columns):
                    col_letter = get_column_letter(i + 1)
                    if ws.column_dimensions[col_letter].width:
                        ws_destino.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                
                # Copiar cabeçalhos com formatação
                for i in range(1, linha_inicio):
                    # Copiar altura da linha
                    if ws.row_dimensions[i].height:
                        ws_destino.row_dimensions[i].height = ws.row_dimensions[i].height
                    
                    for j, cell in enumerate(ws[i], start=1):
                        new_cell = ws_destino.cell(row=i, column=j, value=cell.value)
                        
                        # Copiar formatação
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy(cell.alignment)
                        
                        # Copiar mesclagens
                        if cell.coordinate in ws.merged_cells:
                            for merged_range in ws.merged_cells.ranges:
                                if cell.coordinate in merged_range:
                                    ws_destino.merge_cells(str(merged_range))
                                    break
                
                linha_atual_destino = linha_inicio
                primeira = False
            else:
                linha_atual_destino = ws_destino.max_row + 1

            # Copiar dados com formatação
            for i in range(linha_inicio, ws.max_row + 1):
                if all(cell.value is None for cell in ws[i]):
                    continue
                    
                # Copiar altura da linha
                if ws.row_dimensions[i].height:
                    ws_destino.row_dimensions[linha_atual_destino].height = ws.row_dimensions[i].height
                    
                for j, cell in enumerate(ws[i], start=1):
                    new_cell = ws_destino.cell(row=linha_atual_destino, column=j, value=cell.value)
                    
                    # Copiar formatação
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = copy(cell.alignment)
                    
                    # Copiar mesclagens
                    if cell.coordinate in ws.merged_cells:
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                # Ajustar a referência da mesclagem à nova posição
                                min_col, min_row, max_col, max_row = merged_range.bounds
                                row_offset = linha_atual_destino - i
                                new_range = get_column_letter(min_col) + str(min_row + row_offset) + ":" + get_column_letter(max_col) + str(max_row + row_offset)
                                ws_destino.merge_cells(new_range)
                                break
                    
                linha_atual_destino += 1

        hoje = date.today()
        dia_formatado = hoje.strftime("%d%m%y")
        try:
            nome_arquivo = f"Compilado_{dia_formatado}.xlsx"
            saida_path = os.path.join(CONFIG['COMPILADO_FOLDER'], nome_arquivo)
            wb_compilado.save(saida_path)
        except Exception as e:
            print(f"Erro ao compilar planilhas: {e}")